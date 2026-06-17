import csv
import io
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import select

import models
from database import get_db

router = APIRouter()


@router.get("/reports/summary")
def reports_summary(db: Session = Depends(get_db)):
    cards = db.query(models.Card).all()
    leads = db.query(models.Lead).all()
    total_cards  = len(cards)
    total_value  = sum(c.price or 0 for c in cards)
    total_leads  = len(leads)
    converted    = sum(1 for l in leads if l.converted)
    conv_rate    = round(converted / total_leads * 100, 1) if total_leads else 0
    avg_value    = round(total_value / total_cards, 2) if total_cards else 0

    won_stages = db.query(models.Stage).filter(
        models.Stage.name.ilike('%ganho%') | models.Stage.name.ilike('%won%') |
        models.Stage.name.ilike('%conclu%') | models.Stage.name.ilike('%sucesso%')
    ).all()
    won_ids   = {s.id for s in won_stages}
    won_cards = [c for c in cards if c.stage_id in won_ids]
    won_value = sum(c.price or 0 for c in won_cards)
    win_rate  = round(len(won_cards) / total_cards * 100, 1) if total_cards else 0

    return {
        "total_cards": total_cards,
        "total_value": total_value,
        "total_leads": total_leads,
        "leads_converted": converted,
        "lead_conversion_rate": conv_rate,
        "avg_deal_value": avg_value,
        "won_cards": len(won_cards),
        "won_value": won_value,
        "win_rate": win_rate,
    }


@router.get("/reports/funnel")
def reports_funnel(db: Session = Depends(get_db)):
    pipelines = db.query(models.Pipeline).all()
    result = []
    for p in pipelines:
        stages_data = []
        for s in sorted(p.stages, key=lambda x: x.order):
            stage_cards = db.query(models.Card).filter(models.Card.stage_id == s.id).all()
            stages_data.append({
                "stage_id": s.id,
                "stage_name": s.name,
                "color": s.color,
                "count": len(stage_cards),
                "value": sum(c.price or 0 for c in stage_cards),
            })
        result.append({
            "pipeline_id": p.id,
            "pipeline_name": p.name,
            "stages": stages_data,
        })
    return result


@router.get("/reports/by-source")
def reports_by_source(db: Session = Depends(get_db)):
    cards = db.query(models.Card).all()
    leads = db.query(models.Lead).all()
    source_map = {}
    for c in cards:
        src = c.source or "Não informado"
        if src not in source_map:
            source_map[src] = {"source": src, "cards": 0, "cards_value": 0, "leads": 0}
        source_map[src]["cards"] += 1
        source_map[src]["cards_value"] += c.price or 0
    for l in leads:
        src = l.source or "Não informado"
        if src not in source_map:
            source_map[src] = {"source": src, "cards": 0, "cards_value": 0, "leads": 0}
        source_map[src]["leads"] += 1
    return sorted(source_map.values(), key=lambda x: x["cards"] + x["leads"], reverse=True)


@router.get("/reports/timeline")
def reports_timeline(db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc)
    result = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=1) * (i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i == 0:
            month_end = now
        else:
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(seconds=1)
        cards = db.query(models.Card).filter(
            models.Card.created_at >= month_start,
            models.Card.created_at <= month_end
        ).all()
        leads = db.query(models.Lead).filter(
            models.Lead.created_at >= month_start,
            models.Lead.created_at <= month_end
        ).all() if hasattr(models.Lead, 'created_at') else []
        result.append({
            "month": month_start.strftime("%b/%y"),
            "cards": len(cards),
            "leads": len(leads) if leads else 0,
            "value": sum(c.price or 0 for c in cards),
        })
    return result


@router.get("/reports/by-responsible")
def reports_by_responsible(db: Session = Depends(get_db)):
    users = db.query(models.User).all()
    result = []
    for u in users:
        stmt = select(models.Card).join(models.card_users).where(models.card_users.c.user_id == u.id)
        user_cards = db.execute(stmt).scalars().unique().all()
        total_value = sum(c.price or 0 for c in user_cards)
        result.append({
            "user_id": u.id,
            "user_name": u.name,
            "cards": len(user_cards),
            "value": total_value,
        })
    return sorted(result, key=lambda x: x["value"], reverse=True)


@router.get("/notifications")
def get_notifications(db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc)
    today_str = now.date().isoformat()
    seven_days_ago = now - timedelta(days=7)

    notifications = []

    overdue_tasks = (
        db.query(models.Task)
        .filter(
            models.Task.done == False,
            models.Task.due_date != None,
            models.Task.due_date < today_str,
        )
        .all()
    )
    for task in overdue_tasks:
        notifications.append({
            "id": f"task-{task.id}",
            "type": "overdue_task",
            "title": "Tarefa vencida",
            "body": task.title,
            "card_id": task.card_id,
            "severity": "danger",
            "created_at": task.due_date,
        })

    stalled_cards = (
        db.query(models.Card)
        .filter(
            models.Card.updated_at != None,
            models.Card.updated_at < seven_days_ago,
            models.Card.created_at < seven_days_ago,
        )
        .all()
    )
    for card in stalled_cards:
        notifications.append({
            "id": f"card-{card.id}",
            "type": "stalled_card",
            "title": "Negócio parado",
            "body": f"{card.title} — sem movimentação há mais de 7 dias",
            "card_id": card.id,
            "severity": "warning",
            "created_at": card.updated_at.isoformat(),
        })

    mentions = (
        db.query(models.Activity)
        .filter(
            models.Activity.type == 'note',
            models.Activity.content.contains('@'),
            models.Activity.created_at > seven_days_ago,
        )
        .all()
    )
    for activity in mentions:
        notifications.append({
            "id": f"mention-{activity.id}",
            "type": "mention",
            "title": "Menção em nota",
            "body": activity.content[:80],
            "card_id": activity.card_id,
            "severity": "info",
            "created_at": activity.created_at.isoformat(),
        })

    notifications.sort(key=lambda n: n["created_at"] or "", reverse=True)
    return notifications[:50]


@router.get("/audit-log")
def get_audit_log(
    entity_type: str = "",
    entity_id: int = 0,
    action: str = "",
    actor: str = "",
    date_from: str = "",
    date_to: str = "",
    search: str = "",
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    from datetime import datetime as _dt
    q = db.query(models.AuditLog)
    if entity_type: q = q.filter(models.AuditLog.entity_type == entity_type)
    if entity_id: q = q.filter(models.AuditLog.entity_id == entity_id)
    if action: q = q.filter(models.AuditLog.action == action)
    if actor: q = q.filter(models.AuditLog.actor.ilike(f"%{actor}%"))
    if search: q = q.filter(
        models.AuditLog.entity_name.ilike(f"%{search}%") |
        models.AuditLog.actor.ilike(f"%{search}%") |
        models.AuditLog.details.ilike(f"%{search}%")
    )
    if date_from:
        try:
            q = q.filter(models.AuditLog.created_at >= _dt.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(models.AuditLog.created_at <= _dt.fromisoformat(date_to + "T23:59:59"))
        except Exception:
            pass
    total = q.count()
    items = q.order_by(models.AuditLog.created_at.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "items": [
            {
                "id": i.id,
                "action": i.action,
                "entity_type": i.entity_type,
                "entity_id": i.entity_id,
                "entity_name": i.entity_name,
                "actor": i.actor,
                "actor_email": i.actor_email,
                "details": i.details,
                "created_at": i.created_at.isoformat() if i.created_at else None,
            }
            for i in items
        ]
    }


# ── Available native fields per entity ─────────────────────────────────────
_NATIVE_FIELDS = {
    "cards": [
        {"key": "id",            "label": "ID"},
        {"key": "title",         "label": "Título"},
        {"key": "price",         "label": "Valor (R$)"},
        {"key": "stage_name",    "label": "Etapa"},
        {"key": "pipeline_name", "label": "Funil"},
        {"key": "source",        "label": "Fonte"},
        {"key": "responsible",   "label": "Responsáveis"},
        {"key": "contacts",      "label": "Contatos"},
        {"key": "deal_type",     "label": "Tipo de negócio"},
        {"key": "description",   "label": "Descrição"},
        {"key": "start_date",    "label": "Data início"},
        {"key": "utm_source",    "label": "UTM Source"},
        {"key": "utm_medium",    "label": "UTM Medium"},
        {"key": "utm_campaign",  "label": "UTM Campaign"},
        {"key": "created_at",    "label": "Criado em"},
    ],
    "leads": [
        {"key": "id",           "label": "ID"},
        {"key": "title",        "label": "Título"},
        {"key": "first_name",   "label": "Nome"},
        {"key": "last_name",    "label": "Sobrenome"},
        {"key": "email",        "label": "Email"},
        {"key": "phone",        "label": "Telefone"},
        {"key": "company_name", "label": "Empresa"},
        {"key": "source",       "label": "Fonte"},
        {"key": "stage_name",   "label": "Etapa"},
        {"key": "converted",    "label": "Convertido"},
        {"key": "created_at",   "label": "Criado em"},
    ],
    "contacts": [
        {"key": "id",           "label": "ID"},
        {"key": "first_name",   "label": "Nome"},
        {"key": "last_name",    "label": "Sobrenome"},
        {"key": "email",        "label": "Email"},
        {"key": "phone",        "label": "Telefone"},
        {"key": "company_name", "label": "Empresa"},
        {"key": "position",     "label": "Cargo"},
        {"key": "source",       "label": "Fonte"},
        {"key": "created_at",   "label": "Criado em"},
    ],
    "companies": [
        {"key": "id",        "label": "ID"},
        {"key": "name",      "label": "Nome"},
        {"key": "phone",     "label": "Telefone"},
        {"key": "email",     "label": "Email"},
        {"key": "website",   "label": "Site"},
        {"key": "industry",  "label": "Setor"},
        {"key": "employees", "label": "Funcionários"},
        {"key": "created_at","label": "Criado em"},
    ],
}

_ENTITY_CF_MAP = {"cards": "deal", "leads": "lead", "contacts": "contact", "companies": "company"}

# Default native columns selected when user hasn't customised
_DEFAULT_COLS = {
    "cards":     ["id", "title", "price", "stage_name", "pipeline_name", "source", "responsible", "contacts", "created_at"],
    "leads":     ["id", "title", "first_name", "last_name", "email", "phone", "company_name", "source", "stage_name", "converted", "created_at"],
    "contacts":  ["id", "first_name", "last_name", "email", "phone", "company_name", "position", "source", "created_at"],
    "companies": ["id", "name", "phone", "email", "website", "industry", "employees", "created_at"],
}


@router.get("/reports/export-fields")
def export_fields(entity: str = "cards", db: Session = Depends(get_db)):
    native = _NATIVE_FIELDS.get(entity, [])
    cf_entity = _ENTITY_CF_MAP.get(entity, entity)
    cfs = (db.query(models.CustomField)
           .filter(models.CustomField.entity == cf_entity)
           .order_by(models.CustomField.order, models.CustomField.id)
           .all())
    custom = [{"key": f"cf:{cf.id}", "label": cf.name, "field_type": cf.field_type} for cf in cfs]
    defaults = _DEFAULT_COLS.get(entity, [f["key"] for f in native])
    return {"native": native, "custom": custom, "defaults": defaults}


@router.get("/reports/export")
def export_data(
    entity: str = Query("cards", pattern="^(cards|leads|contacts|companies)$"),
    fmt:    str = Query("csv",   pattern="^(csv|xlsx)$"),
    pipeline_id: int = 0,
    stage_id:    int = 0,
    columns: str = "",   # comma-separated: native keys + "cf:{id}" for custom fields
    db: Session = Depends(get_db),
):
    def fmt_dt(val):
        if not val:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y %H:%M")
        return str(val)

    def fmt_money(val):
        return f"{val:.2f}".replace(".", ",") if val else "0,00"

    # ── Parse requested columns ────────────────────────────────────────────────
    if columns:
        col_list = [c.strip() for c in columns.split(",") if c.strip()]
    else:
        col_list = list(_DEFAULT_COLS.get(entity, [f["key"] for f in _NATIVE_FIELDS.get(entity, [])]))

    native_keys = [c for c in col_list if not c.startswith("cf:")]
    cf_ids      = [int(c[3:]) for c in col_list if c.startswith("cf:")]

    # ── Build label map for native fields ──────────────────────────────────────
    native_label = {f["key"]: f["label"] for f in _NATIVE_FIELDS.get(entity, [])}

    # ── Fetch custom field definitions ─────────────────────────────────────────
    cf_defs = {}
    if cf_ids:
        for cf in db.query(models.CustomField).filter(models.CustomField.id.in_(cf_ids)).all():
            cf_defs[cf.id] = cf

    # ── Load raw records ───────────────────────────────────────────────────────
    if entity == "cards":
        q = db.query(models.Card)
        if pipeline_id:
            q = q.join(models.Stage).filter(models.Stage.pipeline_id == pipeline_id)
        if stage_id:
            q = q.filter(models.Card.stage_id == stage_id)
        rows_db = q.all()
        stage_map    = {s.id: s for s in db.query(models.Stage).all()}
        pipeline_map = {p.id: p for p in db.query(models.Pipeline).all()}

    elif entity == "leads":
        q = db.query(models.Lead)
        if stage_id:
            q = q.filter(models.Lead.stage_id == stage_id)
        rows_db   = q.all()
        stage_map = {s.id: s for s in db.query(models.Stage).all()}
        pipeline_map = {}

    elif entity == "contacts":
        rows_db      = db.query(models.Contact).all()
        stage_map    = {}
        pipeline_map = {}

    else:  # companies
        rows_db      = db.query(models.Company).all()
        stage_map    = {}
        pipeline_map = {}

    # ── Batch-load custom field values (one query for all rows) ────────────────
    cf_value_map: dict = {}  # {(entity_id, field_id): value}
    if cf_ids and rows_db:
        entity_ids = [obj.id for obj in rows_db]
        cfvs = (db.query(models.CustomFieldValue)
                .filter(models.CustomFieldValue.field_id.in_(cf_ids),
                        models.CustomFieldValue.entity_id.in_(entity_ids))
                .all())
        for cfv in cfvs:
            cf_value_map[(cfv.entity_id, cfv.field_id)] = cfv.value or ""

    # ── Helper: get a native field value from an ORM object ───────────────────
    def get_native(key, obj):
        if key == "id":            return obj.id
        if key == "title":         return getattr(obj, "title", "") or ""
        if key == "price":         return fmt_money(getattr(obj, "price", None))
        if key == "stage_name":
            s = stage_map.get(getattr(obj, "stage_id", None))
            return s.name if s else ""
        if key == "pipeline_name":
            s = stage_map.get(getattr(obj, "stage_id", None))
            p = pipeline_map.get(s.pipeline_id) if s else None
            return p.name if p else ""
        if key == "source":        return getattr(obj, "source", "") or ""
        if key == "responsible":   return ", ".join(u.name for u in getattr(obj, "users", []) or [])
        if key == "contacts":      return ", ".join(f"{c.first_name} {c.last_name or ''}".strip() for c in getattr(obj, "contacts", []) or [])
        if key == "deal_type":     return getattr(obj, "deal_type", "") or ""
        if key == "description":   return getattr(obj, "description", "") or ""
        if key == "start_date":    return getattr(obj, "start_date", "") or ""
        if key == "utm_source":    return getattr(obj, "utm_source", "") or ""
        if key == "utm_medium":    return getattr(obj, "utm_medium", "") or ""
        if key == "utm_campaign":  return getattr(obj, "utm_campaign", "") or ""
        if key == "first_name":    return getattr(obj, "first_name", "") or ""
        if key == "last_name":     return getattr(obj, "last_name", "") or ""
        if key == "email":         return getattr(obj, "email", "") or ""
        if key == "phone":         return getattr(obj, "phone", "") or ""
        if key == "company_name":  return getattr(obj, "company_name", "") or ""
        if key == "position":      return getattr(obj, "position", "") or ""
        if key == "converted":     return "Sim" if getattr(obj, "converted", False) else "Não"
        if key == "name":          return getattr(obj, "name", "") or ""
        if key == "website":       return getattr(obj, "website", "") or ""
        if key == "industry":      return getattr(obj, "industry", "") or ""
        if key == "employees":     return getattr(obj, "employees", "") or ""
        if key == "created_at":    return fmt_dt(getattr(obj, "created_at", None))
        return ""

    # ── Build headers + rows ───────────────────────────────────────────────────
    headers = [native_label.get(k, k) for k in native_keys]
    headers += [cf_defs[cf_id].name for cf_id in cf_ids if cf_id in cf_defs]

    rows = []
    for obj in rows_db:
        row = [get_native(k, obj) for k in native_keys]
        row += [cf_value_map.get((obj.id, cf_id), "") for cf_id in cf_ids]
        rows.append(row)

    entity_labels = {"cards": "negocios", "leads": "leads", "contacts": "contatos", "companies": "empresas"}
    filename = f"export_{entity_labels[entity]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # ── CSV ─────────────────────────────────────────────────────────────────────
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_ALL)
        writer.writerow(headers)
        writer.writerows(rows)
        buf.seek(0)
        content = "﻿" + buf.getvalue()  # BOM for Excel UTF-8 detection
        return Response(
            content=content.encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    # ── XLSX ────────────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = entity_labels[entity].capitalize()

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin   = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.get("/search")
def global_search(q: str = "", db: Session = Depends(get_db)):
    if len(q.strip()) < 2:
        return {"cards": [], "leads": [], "contacts": [], "companies": [], "total": 0}

    card_rows = (
        db.query(models.Card, models.Stage)
        .join(models.Stage, models.Card.stage_id == models.Stage.id, isouter=True)
        .filter(
            models.Card.title.ilike(f"%{q}%") |
            models.Card.description.ilike(f"%{q}%")
        )
        .limit(5)
        .all()
    )
    cards = [
        {
            "id": card.id,
            "type": "card",
            "title": card.title or "",
            "subtitle": stage.name if stage else "",
            "stage_id": card.stage_id,
            "url": f"#card-{card.id}",
        }
        for card, stage in card_rows
    ]

    lead_rows = (
        db.query(models.Lead)
        .filter(
            models.Lead.title.ilike(f"%{q}%") |
            models.Lead.first_name.ilike(f"%{q}%") |
            models.Lead.last_name.ilike(f"%{q}%") |
            models.Lead.email.ilike(f"%{q}%") |
            models.Lead.phone.ilike(f"%{q}%") |
            models.Lead.company_name.ilike(f"%{q}%")
        )
        .limit(5)
        .all()
    )
    leads = []
    for lead in lead_rows:
        name_parts = [p for p in [lead.first_name, lead.last_name] if p]
        title = lead.title or (" ".join(name_parts) if name_parts else f"Lead #{lead.id}")
        subtitle = lead.email or lead.phone or lead.company_name or ""
        leads.append({
            "id": lead.id,
            "type": "lead",
            "title": title,
            "subtitle": subtitle,
            "stage_id": lead.stage_id,
            "url": f"#lead-{lead.id}",
        })

    contact_rows = (
        db.query(models.Contact)
        .filter(
            models.Contact.first_name.ilike(f"%{q}%") |
            models.Contact.last_name.ilike(f"%{q}%") |
            models.Contact.email.ilike(f"%{q}%") |
            models.Contact.phone.ilike(f"%{q}%") |
            models.Contact.company_name.ilike(f"%{q}%")
        )
        .limit(5)
        .all()
    )
    contacts = []
    for contact in contact_rows:
        name_parts = [p for p in [contact.first_name, contact.last_name] if p]
        title = " ".join(name_parts) if name_parts else f"Contato #{contact.id}"
        subtitle = contact.email or contact.phone or ""
        contacts.append({
            "id": contact.id,
            "type": "contact",
            "title": title,
            "subtitle": subtitle,
            "url": f"#contact-{contact.id}",
        })

    company_rows = (
        db.query(models.Company)
        .filter(
            models.Company.name.ilike(f"%{q}%") |
            models.Company.company_number.ilike(f"%{q}%")
        )
        .limit(5)
        .all()
    )
    companies = [
        {
            "id": company.id,
            "type": "company",
            "title": company.name or "",
            "subtitle": company.company_number or "",
            "url": f"#company-{company.id}",
        }
        for company in company_rows
    ]

    total = len(cards) + len(leads) + len(contacts) + len(companies)
    return {"cards": cards, "leads": leads, "contacts": contacts, "companies": companies, "total": total}
